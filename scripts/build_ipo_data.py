#!/usr/bin/env python3
"""Fetch 2026 A-share IPOs (主板 / 创业板 / 科创板), sort by listing date, fill codes and issuance/float values."""

from __future__ import annotations

import json
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
AS_OF = datetime(2026, 8, 24)
UA = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://data.eastmoney.com/xg/xg.html",
}


def get_json(url: str, retries: int = 4) -> dict:
    last_err: Exception | None = None
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            time.sleep(1.5 * (i + 1))
    raise RuntimeError(f"GET failed: {url}") from last_err


def board_of(code: str) -> str | None:
    if code.startswith("688"):
        return "科创板"
    if code.startswith(("300", "301")):
        return "创业板"
    if code.startswith(("600", "601", "603", "605")):
        return "沪市主板"
    if code.startswith(("000", "001", "002", "003")):
        return "深市主板"
    return None


def exchange_of(code: str) -> str:
    if code.startswith(("6", "9")):
        return "上交所"
    return "深交所"


def parse_date(value: str | None) -> str | None:
    if not value:
        return None
    return str(value)[:10]


def num(value) -> float | None:
    if value is None or value == "" or value == "-":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def round_or_none(value: float | None, digits: int) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def fetch_ipo_pages(filter_expr: str) -> list[dict]:
    items: list[dict] = []
    for page in range(1, 12):
        params = {
            "sortColumns": "LISTING_DATE,SECURITY_CODE",
            "sortTypes": "-1,-1",
            "pageSize": "50",
            "pageNumber": str(page),
            "reportName": "RPTA_APP_IPOAPPLY",
            "columns": "ALL",
            "source": "WEB",
            "client": "WEB",
            "filter": filter_expr,
        }
        url = "https://datacenter-web.eastmoney.com/api/data/v1/get?" + urllib.parse.urlencode(params)
        payload = get_json(url)
        result = payload.get("result") or {}
        chunk = result.get("data") or []
        items.extend(chunk)
        pages = int(result.get("pages") or 1)
        if page >= pages or not chunk:
            break
        time.sleep(0.25)
    return items


def sina_symbol(code: str) -> str:
    return f"{'sh' if code.startswith('6') else 'sz'}{code}"


def fetch_day_bars(code: str, datalen: int = 320) -> list[dict]:
    """未复权日K：day / open / high / low / close。"""
    url = (
        "https://money.finance.sina.com.cn/quotes_service/api/json_v2.php/"
        f"CN_MarketData.getKLineData?symbol={sina_symbol(code)}&scale=240&ma=no&datalen={datalen}"
    )
    last_err: Exception | None = None
    for i in range(4):
        try:
            req = urllib.request.Request(
                url,
                headers={
                    "User-Agent": UA["User-Agent"],
                    "Referer": "https://finance.sina.com.cn/",
                },
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8")
            rows = json.loads(raw)
            if not isinstance(rows, list):
                return []
            out: list[dict] = []
            for row in rows:
                out.append(
                    {
                        "day": row.get("day"),
                        "open": num(row.get("open")),
                        "high": num(row.get("high")),
                        "low": num(row.get("low")),
                        "close": num(row.get("close")),
                    }
                )
            return out
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            time.sleep(1.2 * (i + 1))
    print(f"warn: day bars failed for {code}: {last_err}")
    return []


def first_two_sessions(
    bars: list[dict], listing_date: str | None
) -> tuple[dict | None, dict | None]:
    if not listing_date or not bars:
        return None, None
    idx = next((i for i, bar in enumerate(bars) if bar.get("day") == listing_date), None)
    if idx is None:
        # 若上市日休市/对不齐，取上市日及之后第一根
        idx = next((i for i, bar in enumerate(bars) if (bar.get("day") or "") >= listing_date), None)
    if idx is None:
        return None, None
    day1 = bars[idx]
    day2 = bars[idx + 1] if idx + 1 < len(bars) else None
    return day1, day2


def listing_status(listing_date: str | None, issue_price: float | None, apply_date: str | None) -> str:
    today = AS_OF.strftime("%Y-%m-%d")
    if listing_date:
        return "已上市" if listing_date <= today else "待上市"
    if apply_date and apply_date > today:
        return "待申购"
    if issue_price:
        return "已发行待上市"
    return "待发行"


def main() -> None:
    by_code: dict[str, dict] = {}
    for filt in ("(LISTING_DATE>='2026-01-01')", "(APPLY_DATE>='2026-01-01')"):
        for row in fetch_ipo_pages(filt):
            code = row.get("SECURITY_CODE")
            if code:
                by_code[code] = row

    stocks = [row for code, row in by_code.items() if board_of(code)]

    day_bars: dict[str, list[dict]] = {}
    for row in stocks:
        code = row["SECURITY_CODE"]
        listing_date = parse_date(row.get("LISTING_DATE"))
        if not listing_date or listing_date > AS_OF.strftime("%Y-%m-%d"):
            continue
        day_bars[code] = fetch_day_bars(code)
        time.sleep(0.15)

    records: list[dict] = []
    for row in stocks:
        code = row["SECURITY_CODE"]
        board = board_of(code)
        listing_date = parse_date(row.get("LISTING_DATE"))
        apply_date = parse_date(row.get("APPLY_DATE"))
        issue_price = num(row.get("ISSUE_PRICE"))
        issue_num_wan = num(row.get("ISSUE_NUM"))  # 万股
        online_shares = num(row.get("ONLINE_ISSUE_NUM"))  # 股
        total_shares = num(row.get("TOTAL_SHARES"))  # 股
        if issue_price:
            raise_yi = num(row.get("DEC_SUMFINA")) or num(row.get("TOTAL_RAISE_FUNDS"))
            if raise_yi is None and issue_num_wan:
                raise_yi = issue_price * issue_num_wan / 10000.0  # 万股*元 / 10000 = 亿元
        else:
            raise_yi = num(row.get("PREDICT_RAISE_FUNDS"))

        online_wan = online_shares / 10000.0 if online_shares else None
        total_shares_wan = total_shares / 10000.0 if total_shares else None
        issue_market_yi = (
            issue_price * total_shares / 1e8 if issue_price and total_shares else None
        )
        # 发行流通值：网上发行股份按发行价计算的市值（上市首日公众可流通部分的发行口径）
        issue_float_yi = (
            issue_price * online_shares / 1e8 if issue_price and online_shares else None
        )
        if issue_float_yi is None and issue_price and issue_num_wan:
            issue_float_yi = issue_price * issue_num_wan / 10000.0

        first_open = num(row.get("OPEN_PRICE"))
        first_close = num(row.get("CLOSE_PRICE"))
        day1, day2 = first_two_sessions(day_bars.get(code) or [], listing_date)
        if day1:
            first_open = day1.get("open") or first_open
            first_close = day1.get("close") or first_close
            first_high = day1.get("high")
            first_low = day1.get("low")
        else:
            high_chg = num(row.get("LD_HIGH_CHANG"))
            first_high = (
                issue_price * (1 + high_chg / 100.0)
                if issue_price and high_chg is not None
                else None
            )
            if first_high is None:
                day_prices = [p for p in (first_open, first_close) if p is not None]
                first_high = max(day_prices) if day_prices else None
            first_low = min(p for p in (first_open, first_close) if p is not None) if any(
                p is not None for p in (first_open, first_close)
            ) else None
        second_high = day2.get("high") if day2 else None
        second_low = day2.get("low") if day2 else None
        second_date = day2.get("day") if day2 else None

        status = listing_status(listing_date, issue_price, apply_date)
        name = row.get("SECURITY_NAME_ABBR") or row.get("SECURITY_NAME")
        records.append(
            {
                "股票代码": code,
                "股票简称": name,
                "板块": board,
                "交易所": exchange_of(code),
                "上市状态": status,
                "上市日期": listing_date,
                "申购日期": apply_date,
                "次日日期": second_date,
                "行业": row.get("INDUSTRY_NAME"),
                "发行价": round_or_none(issue_price, 2),
                "发行总量_万股": round_or_none(issue_num_wan, 2),
                "网上发行_万股": round_or_none(online_wan, 2),
                "发行后总股本_万股": round_or_none(total_shares_wan, 2),
                "募集资金_亿元": round_or_none(raise_yi, 4),
                "发行后总市值_亿元": round_or_none(issue_market_yi, 4),
                "发行流通值_亿元": round_or_none(issue_float_yi, 4),
                "首日开盘价": round_or_none(first_open, 2),
                "首日收盘价": round_or_none(first_close, 2),
                "首日最高价": round_or_none(first_high, 2),
                "首日最低价": round_or_none(first_low, 2),
                "次日最高价": round_or_none(second_high, 2),
                "次日最低价": round_or_none(second_low, 2),
                "首日涨跌幅_pct": round_or_none(num(row.get("LD_CLOSE_CHANGE")), 2),
                "发行市盈率": round_or_none(num(row.get("AFTER_ISSUE_PE")), 2),
                "保荐机构": row.get("RECOMMEND_ORG"),
                "主营业务": row.get("MAIN_BUSINESS"),
                "公司全称": row.get("SECURITY_NAME_FULL"),
            }
        )

    def sort_key(item: dict):
        listing = item["上市日期"] or "9999-12-31"
        apply = item["申购日期"] or "9999-12-31"
        return (listing, apply, item["股票代码"])

    records.sort(key=sort_key)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "asOf": AS_OF.strftime("%Y-%m-%d"),
        "source": "东方财富新股申购接口 RPTA_APP_IPOAPPLY + 新浪日K（首日/次日高低价）",
        "note": "发行流通值 = 网上发行股数 × 发行价（亿元）。无网上发行数据时回退为发行总量 × 发行价。北交所已排除。不含实时行情字段。",
        "count": len(records),
        "boards": {
            "沪市主板": sum(1 for r in records if r["板块"] == "沪市主板"),
            "深市主板": sum(1 for r in records if r["板块"] == "深市主板"),
            "创业板": sum(1 for r in records if r["板块"] == "创业板"),
            "科创板": sum(1 for r in records if r["板块"] == "科创板"),
        },
        "items": records,
    }
    json_path = DATA_DIR / "ipo_2026.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows  # noqa: F401  kept unused-safe

    headers = [
        "序号",
        "股票代码",
        "股票简称",
        "板块",
        "交易所",
        "上市状态",
        "上市日期",
        "申购日期",
        "行业",
        "发行价(元)",
        "发行总量(万股)",
        "网上发行(万股)",
        "发行后总股本(万股)",
        "募集资金(亿元)",
        "发行后总市值(亿元)",
        "发行流通值(亿元)",
        "首日开盘价(元)",
        "首日收盘价(元)",
        "首日最高价(元)",
        "首日最低价(元)",
        "次日日期",
        "次日最高价(元)",
        "次日最低价(元)",
        "首日涨跌幅(%)",
        "发行市盈率",
        "保荐机构",
        "主营业务",
        "公司全称",
    ]
    key_map = {
        "发行价(元)": "发行价",
        "发行总量(万股)": "发行总量_万股",
        "网上发行(万股)": "网上发行_万股",
        "发行后总股本(万股)": "发行后总股本_万股",
        "募集资金(亿元)": "募集资金_亿元",
        "发行后总市值(亿元)": "发行后总市值_亿元",
        "发行流通值(亿元)": "发行流通值_亿元",
        "首日开盘价(元)": "首日开盘价",
        "首日收盘价(元)": "首日收盘价",
        "首日最高价(元)": "首日最高价",
        "首日最低价(元)": "首日最低价",
        "次日最高价(元)": "次日最高价",
        "次日最低价(元)": "次日最低价",
        "首日涨跌幅(%)": "首日涨跌幅_pct",
        "发行市盈率": "发行市盈率",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "2026新股"

    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="微软雅黑", size=10)
    title_fill = PatternFill("solid", fgColor="0F2744")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="F4F7FB")
    listed_fill = PatternFill("solid", fgColor="E8F5E9")
    pending_fill = PatternFill("solid", fgColor="FFF8E1")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = (
        f"2026年主板 / 创业板 / 科创板新股（按上市日期排序）  数据截至 {AS_OF.strftime('%Y-%m-%d')}  "
        f"共 {len(records)} 只  发行流通值=网上发行×发行价"
    )
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    board_counts = payload["boards"]
    ws["A2"] = (
        " · ".join(f"{k} {v}只" for k, v in board_counts.items())
        + "  |  已排除北交所  |  来源：东方财富"
    )
    ws["A2"].font = Font(name="微软雅黑", size=10, color="334155")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[3].height = 32
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(records)}"
    ws.freeze_panes = "A4"

    for idx, item in enumerate(records, 1):
        r = idx + 3
        values = []
        for header in headers:
            if header == "序号":
                values.append(idx)
            elif header in key_map:
                values.append(item.get(key_map[header]))
            else:
                values.append(item.get(header))
        fill = listed_fill if item["上市状态"] == "已上市" else pending_fill
        if idx % 2 == 0 and item["上市状态"] == "已上市":
            fill = alt_fill
        for col, value in enumerate(values, 1):
            cell = ws.cell(r, col, value)
            cell.font = cell_font
            cell.border = thin
            cell.fill = fill
            if col in (1, 2, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(value, float):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if "价" in headers[col - 1] or headers[col - 1] in ("发行市盈率",):
                    cell.number_format = "0.00"
                elif "%" in headers[col - 1]:
                    cell.number_format = "0.00"
                elif "万股" in headers[col - 1]:
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0.0000"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    widths = {
        "A": 6,
        "B": 12,
        "C": 14,
        "D": 12,
        "E": 10,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 18,
        "J": 12,
        "K": 16,
        "L": 14,
        "M": 18,
        "N": 16,
        "O": 18,
        "P": 16,
        "Q": 18,
        "R": 18,
        "S": 16,
        "T": 12,
        "U": 12,
        "V": 14,
        "W": 14,
        "X": 14,
        "Y": 12,
        "Z": 28,
        "AA": 40,
        "AB": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    note = wb.create_sheet("字段说明")
    notes = [
        ["字段", "含义"],
        ["股票代码", "A股证券代码，按板块规则补齐：688科创板 / 301创业板 / 60x沪市主板 / 00x深市主板"],
        ["上市日期", "交易所上市交易日；尚未公布则为空，并按申购日期排在已上市记录之后"],
        ["发行价", "新股发行价格（元/股）"],
        ["发行总量", "本次公开发行股份数量（万股）"],
        ["网上发行", "网上定价发行数量（万股），即面向公众投资者、上市首日通常可流通的部分"],
        ["募集资金", "实际募集资金总额（亿元），含超额配售（如有）"],
        ["发行后总市值", "发行价 × 发行后总股本"],
        ["发行流通值", "网上发行股数 × 发行价（亿元）。衡量发行口径下的可流通市值"],
        ["范围", "仅含2026年主板、创业板、科创板，已排除北交所"],
        ["数据来源", "东方财富 datacenter-web RPTA_APP_IPOAPPLY、新浪日K"],
        ["截止日期", AS_OF.strftime("%Y-%m-%d")],
    ]
    note["A1"].font = Font(name="微软雅黑", size=12, bold=True)
    for r, pair in enumerate(notes, 1):
        note.cell(r, 1, pair[0]).font = Font(name="微软雅黑", bold=r == 1)
        note.cell(r, 2, pair[1]).font = Font(name="微软雅黑")
        note.cell(r, 1).alignment = Alignment(vertical="center")
        note.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
        note.row_dimensions[r].height = 22 if r == 1 else 32
    note.column_dimensions["A"].width = 16
    note.column_dimensions["B"].width = 88

    xlsx_path = DATA_DIR / "2026新股_按上市日期.xlsx"
    wb.save(xlsx_path)
    print(f"wrote {json_path} ({len(records)} rows)")
    print(f"wrote {xlsx_path}")
    print("boards", payload["boards"])


if __name__ == "__main__":
    main()
